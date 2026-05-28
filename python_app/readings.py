import sys
import openpyxl
from datetime import datetime, date
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox, QMessageBox,
    QDialog, QComboBox, QLineEdit, QFileDialog, QFrame, QAbstractItemView, QTextEdit
)
from PySide6.QtCore import Qt, Slot

def get_widget(parent, fc, utils, be, admin_ctx) -> QWidget:
    return ReadingsWidget(parent, fc, utils, be, admin_ctx)

class ReadingsWidget(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30); layout.setSpacing(20)

        header = QHBoxLayout()
        title = QLabel("METER READINGS & QUERIES"); title.setObjectName("page_title")
        header.addWidget(title)
        refresh_btn = QPushButton("🔄 Refresh Readings")
        refresh_btn.clicked.connect(self.refresh_all); header.addWidget(refresh_btn, 0, Qt.AlignRight)
        layout.addLayout(header)

        self.tabs = QTabWidget(); layout.addWidget(self.tabs)
        self.view_tab = ViewReadingsTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.pending_tab = PendingQueriesTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.log_tab = AllQueriesLogTab(self, self.fc, self.utils, self.be, self.admin_ctx)

        self.tabs.addTab(self.view_tab, "📖 View Readings")
        self.tabs.addTab(self.pending_tab, "⚠️ Pending Corrections")
        self.tabs.addTab(self.log_tab, "📜 All Queries Log")

    def refresh_all(self):
        self.view_tab.perform_search(); self.pending_tab.refresh_pending()

class ViewReadingsTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent); self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        f_lay = QHBoxLayout()
        self.f_cycle = QComboBox(); self.f_cycle.setFixedWidth(150)
        self.f_zone = QComboBox(); self.f_zone.addItems(["All"] + [str(z) for z in self.utils.ZONE_RANGE])
        self.f_status = QComboBox(); self.f_status.addItems(["All", "finalized", "skipped", "Anomaly Flagged"])
        btn = QPushButton("🔍 Filter"); btn.clicked.connect(self.perform_search)
        f_lay.addWidget(QLabel("Cycle:")); f_lay.addWidget(self.f_cycle)
        f_lay.addWidget(QLabel("Zone:")); f_lay.addWidget(self.f_zone)
        f_lay.addWidget(QLabel("Status:")); f_lay.addWidget(self.f_status)
        f_lay.addWidget(btn); f_lay.addStretch()
        exp_btn = QPushButton("📤 Export"); exp_btn.clicked.connect(self.do_export)
        imp_btn = QPushButton("📥 Import Corrections"); imp_btn.clicked.connect(self.do_import)
        f_lay.addWidget(exp_btn); f_lay.addWidget(imp_btn); layout.addLayout(f_lay)

        self.table = QTableWidget(0, 9); self.table.setHorizontalHeaderLabels(["Date", "CIN", "Reader", "Prev", "Curr", "Usage", "Bill", "Status", "Edited"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch); self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows); self.table.doubleClicked.connect(self.on_row_double); layout.addWidget(self.table)
        self.utils.run_in_thread(self.fc.list_billing_cycles, callback=lambda cs: [self.f_cycle.clear(), self.f_cycle.addItems(["All"] + [c["cycle_id"] for c in cs])])

    def perform_search(self):
        cid, zone, stat = self.f_cycle.currentText(), self.f_zone.currentText(), self.f_status.currentText()
        def fetch():
            readings = self.fc.get_readings_for_cycle(cid if cid != "All" else "")
            filtered = []
            for r in readings:
                c = self.fc.get_consumer(r["cin_no"])
                if not c: continue
                if zone != "All" and int(c.get("zone", 0)) != int(zone): continue
                if stat == "skipped" and r.get("status") != "skipped": continue
                if stat == "finalized" and r.get("status") != "finalized": continue
                if stat == "Anomaly Flagged" and not r.get("anomaly_flagged"): continue
                filtered.append((r, c))
            return filtered
        def done(data):
            self.table.setRowCount(0)
            for r, c in data:
                row = self.table.rowCount(); self.table.insertRow(row)
                bb = r.get("full_bill_breakdown", {})
                items = [self.utils.format_date(r.get("submitted_at")), r["cin_no"], r.get("reader_name", ""), f"{r.get('previous_reading',0):.2f}",
                         f"{r.get('current_reading',0):.2f}" if r.get('current_reading') is not None else "Skipped", f"{r.get('consumption',0):.2f}",
                         self.utils.format_currency(bb.get("total_amount", 0)), r.get("status"), "Yes" if r.get("edited_by_admin") else "No"]
                for i, v in enumerate(items):
                    item = QTableWidgetItem(str(v)); item.setData(Qt.UserRole, r["reading_id"]); self.table.setItem(row, i, item)
        self.utils.run_in_thread(fetch, callback=done)

    def on_row_double(self, index):
        rid = self.table.item(index.row(), 0).data(Qt.UserRole)
        ReadingDetailDialog(self, rid, self.fc, self.utils, self.be, self.admin_ctx, self.perform_search).exec()

    def do_export(self):
        cid = self.f_cycle.currentText()
        if cid == "All": return QMessageBox.warning(self, "Error", "Select a cycle.")
        path, _ = QFileDialog.getSaveFileName(self, "Save Export", f"Readings_{cid}.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        def run():
            rs = self.fc.get_readings_for_cycle(cid); wb = openpyxl.Workbook(); ws = wb.active
            headers = ["reading_id", "cin_no", "cycle_id", "reader_name", "previous_reading", "current_reading", "notes"]
            ws.append(headers)
            for r in rs: ws.append([r.get(h) for h in headers])
            wb.save(path)
        self.utils.run_in_thread(run, callback=lambda _: QMessageBox.information(self, "Done", "Export finished."))

    def do_import(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel", "", "Excel Files (*.xlsx)")
        if not path: return
        def run():
            ws = openpyxl.load_workbook(path, data_only=True).active; rows = list(ws.iter_rows(values_only=True)); headers = [str(h).strip() for h in rows[0]]
            updates = []
            for r in rows[1:]:
                if not any(r): continue
                d = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
                if d.get("reading_id") and d.get("current_reading") is not None:
                    updates.append({"reading_id": d["reading_id"], "current_reading": float(d["current_reading"]), "notes": d.get("notes", "Bulk update")})
            return self.fc.bulk_update_readings(updates, self.admin_ctx["name"])
        self.utils.run_in_thread(run, callback=lambda res: [QMessageBox.information(self, "Done", f"Updated {res['success']} rows."), self.perform_search()])

class ReadingDetailDialog(QDialog):
    def __init__(self, parent, rid, fc, utils, be, admin_ctx, refresh_cb):
        super().__init__(parent); self.rid = rid; self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx; self.refresh_cb = refresh_cb
        self.setWindowTitle(f"Reading Detail — {rid}"); self.resize(700, 600); self.setup_ui()

    def setup_ui(self):
        self.lay = QVBoxLayout(self); self.info = QLabel("Loading..."); self.lay.addWidget(self.info)
        def fetch():
            r = self.fc.get_reading(self.rid); c = self.fc.get_consumer(r["cin_no"]) if r else None
            return r, c
        def done(res):
            r, c = res; self.r = r; self.c = c
            if not r: return self.info.setText("Not found.")
            self.info.setText(f"<b>CIN:</b> {r['cin_no']} | <b>Name:</b> {c['name'] if c else 'N/A'} | <b>Reader:</b> {r.get('reader_name')}")
            grid = QGroupBox("Reading Info"); glay = QFormLayout(grid)
            glay.addRow("Status:", QLabel(r.get("status"))); glay.addRow("Prev Reading:", QLabel(f"{r.get('previous_reading',0):.2f} KL"))
            glay.addRow("Curr Reading:", QLabel(f"{r.get('current_reading',0):.2f} KL" if r.get('current_reading') is not None else "Skipped"))
            glay.addRow("Consumption:", QLabel(f"{r.get('consumption',0):.2f} KL")); self.lay.addWidget(grid)

            bb = r.get("full_bill_breakdown", {})
            if bb:
                box = QGroupBox("Bill Breakdown"); blay = QVBoxLayout(box); t = QTextEdit(); t.setReadOnly(True); t.setFont("Courier New")
                txt = f"Water: {self.utils.format_currency(bb.get('water_charge'))}\nFixed: {self.utils.format_currency(bb.get('fixed_charge'))}\nSewer: {self.utils.format_currency(bb.get('sewerage_tax'))}\nSTP:   {self.utils.format_currency(bb.get('stp_charge'))}\nIDS:   {self.utils.format_currency(bb.get('ids_charge'))}\nTotal: {self.utils.format_currency(bb.get('total_amount'))}"
                t.setPlainText(txt); blay.addWidget(t); self.lay.addWidget(box)

            if r.get("status") == "finalized":
                btn = QPushButton("🛠️ Admin Override Edit"); btn.clicked.connect(self.open_override); self.lay.addWidget(btn)
        self.utils.run_in_thread(fetch, callback=done)

    def open_override(self):
        if AdminOverrideDialog(self, self.r, self.c, self.fc, self.utils, self.be, self.admin_ctx).exec():
            self.refresh_cb(); self.accept()

class AdminOverrideDialog(QDialog):
    def __init__(self, parent, r, c, fc, utils, be, admin_ctx):
        super().__init__(parent); self.r = r; self.c = c; self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx
        self.setWindowTitle("Admin Override"); self.resize(800, 500); self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        h = QHBoxLayout(); layout.addLayout(h)

        self.old_table = QTableWidget(6, 2); self.new_table = QTableWidget(6, 2)
        for t in [self.old_table, self.new_table]:
            t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            t.setHorizontalHeaderLabels(["Item", "Value"])
            t.verticalHeader().setVisible(False)

        g1 = QGroupBox("Original Bill"); l1 = QVBoxLayout(g1); l1.addWidget(self.old_table); h.addWidget(g1)
        g2 = QGroupBox("New Simulated Bill"); l2 = QVBoxLayout(g2); l2.addWidget(self.new_table); h.addWidget(g2)
        
        self.f_new = QLineEdit(str(self.r["current_reading"])); self.f_new.textChanged.connect(self.simulate)
        self.f_note = QTextEdit(); self.f_note.setPlaceholderText("Required reason..."); layout.addWidget(QLabel("New Current Reading:")); layout.addWidget(self.f_new); layout.addWidget(QLabel("Reason:")); layout.addWidget(self.f_note)
        btn = QPushButton("✅ Apply Override"); btn.clicked.connect(self.save); layout.addWidget(btn)
        self.simulate()

    def simulate(self):
        try:
            curr = float(self.f_new.text()); prev = float(self.r["previous_reading"]); cons = curr - prev
            if cons < 0: return
            rates = self.fc.get_charges_config(); cycle = self.fc.get_billing_cycle(self.r["cycle_id"])
            lp = self.utils.parse_date(cycle["last_payment_date"]) if cycle else None
            res = self.be.calculate_bill(cons, self.c, rates, previous_outstanding=float(self.c["outstanding_balance"])-self.r["full_bill_breakdown"]["total_amount"], credit_balance=float(self.c["credit_balance"]), last_payment_date=lp, payment_date=date.today())
            self.fill_sim(self.old_table, self.r.get("full_bill_breakdown", {})); self.fill_sim(self.new_table, res)
        except: pass

    def fill_sim(self, t, bb):
        items = [("Usage", f"{bb.get('consumption',0):.2f} KL"), ("Water", self.utils.format_currency(bb.get("water_charge",0))), ("Sewer", self.utils.format_currency(bb.get("sewerage_tax",0))), ("STP", self.utils.format_currency(bb.get("stp_charge",0))), ("IDS", self.utils.format_currency(bb.get("ids_charge",0))), ("TOTAL", self.utils.format_currency(bb.get("total_amount",0)))]
        for i, (k, v) in enumerate(items): t.setItem(i, 0, QTableWidgetItem(k)); t.setItem(i, 1, QTableWidgetItem(v))

    def save(self):
        note = self.f_note.toPlainText().strip()
        if not note or not self.f_new.text(): return QMessageBox.warning(self, "Error", "Fill all fields.")
        def run(): self.fc.admin_update_reading(self.r["reading_id"], {"current_reading": float(self.f_new.text()), "notes": f"Admin: {note}"}, self.admin_ctx["name"])
        self.utils.run_in_thread(run, callback=lambda _: self.accept())

class PendingQueriesTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent); self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx; self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.table = QTableWidget(0, 7); self.table.setHorizontalHeaderLabels(["Date", "CIN", "Name", "Reader", "Submitted", "Requested", "Reason"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch); self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows); self.table.doubleClicked.connect(self.on_row_double); layout.addWidget(self.table)
        self.refresh_pending()

    def refresh_pending(self):
        def fetch(): return self.fc.get_pending_correction_queries()
        def done(qs):
            self.table.setRowCount(0); self.qs = qs
            idx = self.parent().parent().tabs.indexOf(self)
            self.parent().parent().tabs.setTabText(idx, f"⚠️ Pending ({len(qs)})")
            self.parent().parent().tabs.tabBar().setTabTextColor(idx, Qt.red if qs else Qt.black)
            for q in qs:
                row = self.table.rowCount(); self.table.insertRow(row)
                snap = q.get("consumer_info_snapshot", {})
                items = [
                    self.utils.format_date(q.get("created_at")),
                    q["cin_no"],
                    snap.get("name", "Unknown"),
                    q.get("reader_name", "Field Reader"),
                    f"{q.get('submitted_reading', 0.0):.2f}",
                    f"{q.get('requested_corrected_reading', 0.0):.2f}",
                    q.get("reason", "")
                ]
                for i, v in enumerate(items):
                    self.table.setItem(row, i, QTableWidgetItem(str(v)))
        self.utils.run_in_thread(fetch, callback=done)

    def on_row_double(self, index):
        q = self.qs[index.row()]
        dlg = QDialog(self); dlg.setWindowTitle("Review Query"); dlg.resize(500, 400); lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel(f"<b>Query for {q['cin_no']}</b> by {q.get('reader_name')}"))
        lay.addWidget(QLabel(f"Prev: {q.get('previous_reading'):.2f} | Sub: {q.get('submitted_reading'):.2f} | Req: {q.get('requested_corrected_reading'):.2f}"))
        lay.addWidget(QLabel(f"Reason: {q.get('reason')}")); note = QTextEdit(); note.setPlaceholderText("Rejection note (required for reject)"); lay.addWidget(note)
        h = QHBoxLayout(); app_b = QPushButton("✅ Approve"); rej_b = QPushButton("❌ Reject"); h.addWidget(app_b); h.addWidget(rej_b); lay.addLayout(h)
        app_b.clicked.connect(lambda: self.utils.run_in_thread(lambda: self.fc.approve_correction_query(q["query_id"], self.admin_ctx["name"]), callback=lambda _: [dlg.accept(), self.refresh_pending()]))
        rej_b.clicked.connect(lambda: [self.utils.run_in_thread(lambda: self.fc.reject_correction_query(q["query_id"], note.toPlainText(), self.admin_ctx["name"]), callback=lambda _: [dlg.accept(), self.refresh_pending()]) if note.toPlainText().strip() else QMessageBox.warning(dlg, "Error", "Note required.")])
        dlg.exec()

class AllQueriesLogTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent); self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx; self.setup_ui()
    def setup_ui(self):
        layout = QVBoxLayout(self); self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(["Created", "CIN", "Reader", "Submitted", "Requested", "Status", "Resolved"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch); layout.addWidget(self.table)
        self.utils.run_in_thread(self.fc.get_all_correction_queries, callback=self.done)
    def done(self, qs):
        self.table.setRowCount(0)
        for q in qs:
            row = self.table.rowCount(); self.table.insertRow(row)
            items = [
                self.utils.format_date(q.get("created_at")),
                q["cin_no"],
                q.get("reader_name", ""),
                f"{q.get('submitted_reading', 0.0):.2f}",
                f"{q.get('requested_corrected_reading', 0.0):.2f}",
                q.get("status", ""),
                self.utils.format_date(q.get("resolved_at"))
            ]
            for i, v in enumerate(items):
                self.table.setItem(row, i, QTableWidgetItem(str(v)))
