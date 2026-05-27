import sys
import os
import openpyxl
from datetime import datetime, date
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTabWidget,
    QLineEdit, QComboBox, QCheckBox, QGroupBox, QFormLayout, QSpinBox,
    QDoubleSpinBox, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QFileDialog, QProgressBar, QDialog, QScrollArea, QFrame, QAbstractItemView
)
from PySide6.QtCore import Qt, Slot
from firebase_admin import firestore

def get_widget(parent, fc, utils, be, admin_ctx) -> QWidget:
    return ConsumersWidget(parent, fc, utils, be, admin_ctx)

class ConsumersWidget(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc
        self.utils = utils
        self.be = be
        self.admin_ctx = admin_ctx
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        # Header
        header = QHBoxLayout()
        title = QLabel("CONSUMER MANAGEMENT")
        title.setObjectName("page_title")
        header.addWidget(title)

        refresh_btn = QPushButton("🔄 Refresh All")
        refresh_btn.clicked.connect(self.refresh_all)
        header.addWidget(refresh_btn, 0, Qt.AlignRight)
        layout.addLayout(header)

        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self.search_tab = SearchTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.add_tab = AddConsumerTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.import_tab = BulkImportTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.export_tab = ExportTab(self, self.fc, self.utils, self.be, self.admin_ctx)
        self.replace_tab = MeterReplacementTab(self, self.fc, self.utils, self.be, self.admin_ctx)

        self.tabs.addTab(self.search_tab, "🔍 Search / View")
        self.tabs.addTab(self.add_tab, "➕ Add Consumer")
        self.tabs.addTab(self.import_tab, "📥 Bulk Import")
        self.tabs.addTab(self.export_tab, "📤 Export CSD")
        self.tabs.addTab(self.replace_tab, "🔧 Meter Replacement")

    def refresh_all(self):
        self.search_tab.perform_search()

class SearchTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        filter_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search CIN, Name, or Meter Serial...")
        self.search_input.setFixedWidth(300)

        self.zone_cb = QComboBox()
        self.zone_cb.addItems(["All Zones"] + [str(z) for z in self.utils.ZONE_RANGE])

        self.status_cb = QComboBox()
        self.status_cb.addItems(["All Statuses"] + self.utils.CONSUMER_STATUS_OPTIONS)

        search_btn = QPushButton("Search")
        search_btn.clicked.connect(self.perform_search)

        filter_layout.addWidget(self.search_input)
        filter_layout.addWidget(self.zone_cb)
        filter_layout.addWidget(self.status_cb)
        filter_layout.addWidget(search_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels([
            "CIN", "Name", "Zone", "Supply", "Category",
            "Meter Serial", "Last Reading", "Outstanding", "Status"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.doubleClicked.connect(self.on_row_double_click)
        layout.addWidget(self.table)

        self.perform_search()

    def perform_search(self):
        query = self.search_input.text().strip().lower()
        zone_f = self.zone_cb.currentText()
        status_f = self.status_cb.currentText()

        def fetch():
            filters = {"is_active": True}
            if zone_f != "All Zones": filters["zone"] = int(zone_f)
            if status_f != "All Statuses": filters["status"] = status_f
            consumers = self.fc.list_consumers(filters)
            if query:
                consumers = [c for c in consumers if query in c["cin_no"].lower() or query in c["name"].lower() or query in str(c.get("meter_serial_no", "")).lower()]
            return consumers

        def done(consumers):
            self.table.setRowCount(0)
            for c in consumers:
                row = self.table.rowCount()
                self.table.insertRow(row)
                supply = c.get("water_supply_type", "PHED")
                badge = QLabel(supply)
                badge.setAlignment(Qt.AlignCenter)
                color = "#E8913A" if supply == "Own Supply" else "#1A3A6B"
                badge.setStyleSheet(f"background-color: {color}; color: white; border-radius: 10px; padding: 2px; font-weight: bold;")
                
                self.table.setItem(row, 0, QTableWidgetItem(c["cin_no"]))
                self.table.setItem(row, 1, QTableWidgetItem(c["name"]))
                self.table.setItem(row, 2, QTableWidgetItem(str(c["zone"])))
                self.table.setCellWidget(row, 3, badge)
                self.table.setItem(row, 4, QTableWidgetItem(c["category"]))
                self.table.setItem(row, 5, QTableWidgetItem(c.get("meter_serial_no", "")))
                self.table.setItem(row, 6, QTableWidgetItem(f"{c.get('last_reading', 0):.2f} KL"))
                self.table.setItem(row, 7, QTableWidgetItem(self.utils.format_currency(c.get("outstanding_balance", 0))))
                self.table.setItem(row, 8, QTableWidgetItem(c.get("consumer_status", "Active")))

        self.utils.run_in_thread(fetch, callback=done)

    def on_row_double_click(self, index):
        cin_no = self.table.item(index.row(), 0).text()
        ConsumerDetailDialog(self, cin_no, self.fc, self.utils, self.be, self.admin_ctx, self.perform_search).exec()

class AddConsumerTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        container = QWidget(); self.form = QFormLayout(container)
        self.form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        
        self.f_cin = QLineEdit()
        self.f_name = QLineEdit()
        self.f_zone = QComboBox(); self.f_zone.addItems([str(z) for z in self.utils.ZONE_RANGE])
        self.f_cat = QComboBox(); self.f_cat.addItems(self.utils.CATEGORY_OPTIONS)
        
        self.f_supply = QComboBox(); self.f_supply.addItems(self.utils.WATER_SUPPLY_TYPE_OPTIONS)
        self.f_supply.currentTextChanged.connect(self._on_supply_type_changed)
        
        self.f_sewer = QCheckBox("Has Sewer Connection")
        
        self.sew_group = QGroupBox("Own Supply — Sewerage Details"); self.sew_group.setVisible(False)
        sew_lay = QFormLayout(self.sew_group)
        self.f_sub_cat = QComboBox(); self.f_sub_cat.addItems(["-- Select --"] + self.utils.SEWERAGE_SUB_CATEGORY_OPTIONS)
        self.f_sub_cat.currentTextChanged.connect(self._on_sub_category_changed)

        self.f_rooms = QSpinBox(); self.f_rooms.setRange(1, 9999); self.f_rooms_row = QLabel("Rooms:"); self.f_rooms.setVisible(False)
        self.f_plot = QDoubleSpinBox(); self.f_plot.setRange(0.01, 999999); self.f_plot.setDecimals(2); self.f_plot.setVisible(False)

        sew_lay.addRow("Sub-Category:", self.f_sub_cat)
        sew_lay.addRow("Rooms:", self.f_rooms)
        sew_lay.addRow("Plot Area (sqm):", self.f_plot)

        self.f_size = QComboBox(); self.f_size.addItems(self.utils.METER_SIZE_OPTIONS)
        self.f_serial = QLineEdit()
        self.f_init = QDoubleSpinBox(); self.f_init.setRange(0, 999999)
        self.f_phone = QLineEdit()
        self.f_aadhaar = QLineEdit()
        self.f_status = QComboBox(); self.f_status.addItems(self.utils.CONSUMER_STATUS_OPTIONS)

        self.form.addRow("CIN Number <span style='color:#E8913A'>*</span>:", self.f_cin)
        self.form.addRow("Name <span style='color:#E8913A'>*</span>:", self.f_name)
        self.form.addRow("Zone <span style='color:#E8913A'>*</span>:", self.f_zone)
        self.form.addRow("Category <span style='color:#E8913A'>*</span>:", self.f_cat)
        self.form.addRow("Supply Type <span style='color:#E8913A'>*</span>:", self.f_supply)
        self.form.addRow(self.f_sewer)
        self.form.addRow(self.sew_group)
        self.form.addRow("Meter Size:", self.f_size)
        self.form.addRow("Meter Serial:", self.f_serial)
        self.form.addRow("Initial KL:", self.f_init)
        self.form.addRow("Phone:", self.f_phone)
        self.form.addRow("Aadhaar:", self.f_aadhaar)
        self.form.addRow("Status:", self.f_status)

        save_btn = QPushButton("💾 Save Consumer")
        save_btn.clicked.connect(self.save)
        self.form.addRow(save_btn)

        scroll.setWidget(container); layout.addWidget(scroll)

    def _on_supply_type_changed(self):
        is_own = self.f_supply.currentText() == "Own Supply"
        self.sew_group.setVisible(is_own)
        if not is_own:
            self.f_sub_cat.setCurrentIndex(0); self.f_rooms.setVisible(False); self.f_plot.setVisible(False)

    def _on_sub_category_changed(self):
        sub = self.f_sub_cat.currentText()
        self.f_rooms.setVisible(sub == "Hotel")
        self.f_plot.setVisible(sub == "Domestic (Own Supply)")

    def save(self):
        data = {
            "cin_no": self.f_cin.text().strip(), "name": self.f_name.text().strip(),
            "zone": int(self.f_zone.currentText()), "category": self.f_cat.currentText(),
            "water_supply_type": self.f_supply.currentText(), "has_sewer_connection": self.f_sewer.isChecked(),
            "meter_size": self.f_size.currentText(), "meter_serial_no": self.f_serial.text().strip(),
            "initial_meter_reading": self.f_init.value(), "contact_number": self.f_phone.text().strip(),
            "aadhaar_phed_no": self.f_aadhaar.text().strip(), "consumer_status": self.f_status.currentText(),
            "is_active": True, "credit_balance": 0.0, "outstanding_balance": 0.0
        }
        if not data["cin_no"] or not data["name"]:
            return QMessageBox.warning(self, "Required", "CIN and Name are required.")
        if data["water_supply_type"] == "Own Supply":
            sub = self.f_sub_cat.currentText()
            if sub == "-- Select --": return QMessageBox.warning(self, "Required", "Sub-category is required.")
            data["sewerage_sub_category"] = sub
            if sub == "Hotel": data["num_rooms"] = self.f_rooms.value()
            elif sub == "Domestic (Own Supply)": data["plot_area_sqmtr"] = self.f_plot.value()

        self.utils.run_in_thread(lambda: self.fc.create_consumer(data, self.admin_ctx["name"]),
                                 callback=lambda c: [QMessageBox.information(self, "Success", f"Created {c}"), self.parent().parent().search_tab.perform_search()])

class ConsumerDetailDialog(QDialog):
    def __init__(self, parent, cin_no, fc, utils, be, admin_ctx, refresh_cb):
        super().__init__(parent)
        self.cin_no = cin_no; self.fc = fc; self.utils = utils; self.be = be; self.admin_ctx = admin_ctx; self.refresh_cb = refresh_cb
        self.setWindowTitle(f"Consumer Detail — {cin_no}"); self.resize(900, 700); self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget(); layout.addWidget(self.tabs)
        self.profile_tab = QWidget(); self.tabs.addTab(self.profile_tab, "👤 Profile")
        self.readings_tab = QTableWidget(0, 6); self.tabs.addTab(self.readings_tab, "📊 Readings")
        self.payments_tab = QTableWidget(0, 4); self.tabs.addTab(self.payments_tab, "💳 Payments")
        self.adj_tab = QTableWidget(0, 4); self.tabs.addTab(self.adj_tab, "🔧 Adjustments")
        self.rep_tab = QTableWidget(0, 4); self.tabs.addTab(self.rep_tab, "⚙️ Replacements")
        
        for t in [self.readings_tab, self.payments_tab, self.adj_tab, self.rep_tab]:
            t.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        self.readings_tab.setHorizontalHeaderLabels(["Date", "Cycle", "Prev", "Curr", "Usage", "Status"])
        self.payments_tab.setHorizontalHeaderLabels(["Date", "Receipt", "Mode", "Amount"])
        self.adj_tab.setHorizontalHeaderLabels(["Date", "Type", "Reason", "Amount"])
        self.rep_tab.setHorizontalHeaderLabels(["Date", "Old Serial", "New Serial", "Init KL"])
        self.load_data()

    def load_data(self):
        def fetch():
            c = self.fc.get_consumer(self.cin_no)
            r = self.fc.get_readings_for_cycle("", self.cin_no)
            p = self.fc.get_payments_for_consumer(self.cin_no)
            a = self.fc.get_adjustments_for_consumer(self.cin_no)
            m = self.fc.get_meter_replacement_history(self.cin_no)
            return c, r, p, a, m
        def done(res):
            c, readings, payments, adjustments, replacements = res
            self.c = c; self.build_profile(c)
            self.fill_table(self.readings_tab, readings, ["reading_date", "cycle_id", "previous_reading", "current_reading", "consumption", "status"])
            self.fill_table(self.payments_tab, payments, ["payment_date", "receipt_number", "payment_mode", "amount"])
            self.fill_table(self.adj_tab, adjustments, ["applied_at", "type", "reason_note", "amount"])
            self.fill_table(self.rep_tab, replacements, ["replacement_date", "old_meter_serial", "new_meter_serial", "new_initial_reading"])
        self.utils.run_in_thread(fetch, callback=done)

    def fill_table(self, table, data, keys):
        table.setRowCount(0)
        for d in data:
            row = table.rowCount(); table.insertRow(row)
            for i, k in enumerate(keys):
                val = d.get(k, "")
                if k in ["amount", "outstanding"]: val = self.utils.format_currency(val)
                elif "reading" in k or k == "consumption": val = f"{float(val or 0):.2f}"
                elif k == "applied_at": val = self.utils.format_date(val)
                table.setItem(row, i, QTableWidgetItem(str(val)))

    def build_profile(self, c):
        if self.profile_tab.layout():
            QWidget().setLayout(self.profile_tab.layout())
        lay = QVBoxLayout(self.profile_tab)
        g = QGroupBox("Core Profile"); form = QFormLayout(g)
        items = [("CIN", c["cin_no"]), ("Name", c["name"]), ("Zone", str(c["zone"])), ("Supply", c.get("water_supply_type", "PHED")), ("Sewer", "Yes" if c.get("has_sewer_connection") else "No")]
        if c.get("water_supply_type") == "Own Supply":
            items.append(("Sub-Category", c.get("sewerage_sub_category")))
            if "num_rooms" in c: items.append(("Rooms", str(c["num_rooms"])))
            if "plot_area_sqmtr" in c: items.append(("Plot Area", f"{c['plot_area_sqmtr']} sqm"))
        for l, v in items: form.addRow(f"<b>{l}:</b>", QLabel(v))
        lay.addWidget(g)
        
        btns = QHBoxLayout(); edit_b = QPushButton("✏️ Edit"); edit_b.clicked.connect(self.open_edit)
        deact_b = QPushButton("Deactivate" if c.get("is_active") else "Reactivate"); deact_b.clicked.connect(self.toggle_status)
        btns.addWidget(edit_b); btns.addWidget(deact_b); lay.addLayout(btns)

    def open_edit(self):
        if EditConsumerDialog(self, self.c, self.fc, self.utils, self.admin_ctx).exec():
            self.load_data(); self.refresh_cb()

    def toggle_status(self):
        active = self.c.get("is_active", True)
        if QMessageBox.question(self, "Confirm", f"Sure to {'Deactivate' if active else 'Reactivate'}?") == QMessageBox.Yes:
            fn = self.fc.deactivate_consumer if active else self.fc.reactivate_consumer
            self.utils.run_in_thread(lambda: fn(self.cin_no, self.admin_ctx["name"]), callback=lambda _: [self.load_data(), self.refresh_cb()])

class EditConsumerDialog(QDialog):
    def __init__(self, parent, c, fc, utils, admin_ctx):
        super().__init__(parent)
        self.c = c; self.fc = fc; self.utils = utils; self.admin_ctx = admin_ctx
        self.setWindowTitle(f"Edit {c['cin_no']}"); self.resize(600, 700); self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self); scroll = QScrollArea(); scroll.setWidgetResizable(True)
        container = QWidget(); self.form = QFormLayout(container)
        self.f_name = QLineEdit(self.c["name"])
        self.f_supply = QComboBox(); self.f_supply.addItems(self.utils.WATER_SUPPLY_TYPE_OPTIONS); self.f_supply.setCurrentText(self.c.get("water_supply_type", "PHED"))
        self.f_supply.currentTextChanged.connect(self._on_supply_type_changed)
        self.f_sewer = QCheckBox("Has Sewer Connection"); self.f_sewer.setChecked(bool(self.c.get("has_sewer_connection")))
        self.sew_group = QGroupBox("Sewerage Details"); sew_lay = QFormLayout(self.sew_group)
        self.f_sub_cat = QComboBox(); self.f_sub_cat.addItems(self.utils.SEWERAGE_SUB_CATEGORY_OPTIONS); self.f_sub_cat.setCurrentText(self.c.get("sewerage_sub_category", ""))
        self.f_sub_cat.currentTextChanged.connect(self._on_sub_category_changed)
        self.f_rooms = QSpinBox(); self.f_rooms.setRange(1, 9999); self.f_rooms.setValue(self.c.get("num_rooms", 1))
        self.f_plot = QDoubleSpinBox(); self.f_plot.setRange(0.01, 999999); self.f_plot.setValue(self.c.get("plot_area_sqmtr", 0.01))
        sew_lay.addRow("Sub-Category:", self.f_sub_cat); sew_lay.addRow("Rooms:", self.f_rooms); sew_lay.addRow("Plot Area:", self.f_plot)
        self.form.addRow("Name:", self.f_name); self.form.addRow("Supply:", self.f_supply); self.form.addRow(self.f_sewer); self.form.addRow(self.sew_group)
        save_btn = QPushButton("Save"); save_btn.clicked.connect(self.save); self.form.addRow(save_btn)
        scroll.setWidget(container); layout.addWidget(scroll); self._on_supply_type_changed()

    def _on_supply_type_changed(self):
        is_own = self.f_supply.currentText() == "Own Supply"; self.sew_group.setVisible(is_own)
        if is_own: self._on_sub_category_changed()

    def _on_sub_category_changed(self):
        sub = self.f_sub_cat.currentText(); self.f_rooms.setVisible(sub == "Hotel"); self.f_plot.setVisible(sub == "Domestic (Own Supply)")

    def save(self):
        updates = {"name": self.f_name.text().strip(), "water_supply_type": self.f_supply.currentText(), "has_sewer_connection": self.f_sewer.isChecked()}
        if updates["water_supply_type"] == "PHED":
            if self.c.get("water_supply_type") == "Own Supply":
                updates.update({"sewerage_sub_category": firestore.DELETE_FIELD, "num_rooms": firestore.DELETE_FIELD, "plot_area_sqmtr": firestore.DELETE_FIELD})
        else:
            sub = self.f_sub_cat.currentText(); updates["sewerage_sub_category"] = sub
            if sub == "Hotel": updates["num_rooms"] = self.f_rooms.value()
            elif sub == "Domestic (Own Supply)": updates["plot_area_sqmtr"] = self.f_plot.value()
        self.utils.run_in_thread(lambda: self.fc.update_consumer(self.c["cin_no"], updates, self.admin_ctx["name"]), callback=lambda _: self.accept())

class BulkImportTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.admin_ctx = admin_ctx; self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("<b>Bulk Import CSD (Excel)</b>"))
        self.path_edit = QLineEdit(); self.path_edit.setReadOnly(True)
        btn_h = QHBoxLayout(); btn_h.addWidget(self.path_edit)
        browse_b = QPushButton("📂 Browse"); browse_b.clicked.connect(self.browse); btn_h.addWidget(browse_b); layout.addLayout(btn_h)
        self.progress = QProgressBar(); self.progress.setVisible(False); layout.addWidget(self.progress)
        import_b = QPushButton("🚀 Import"); import_b.clicked.connect(self.run_import); layout.addWidget(import_b)
        dl_b = QPushButton("📥 Download Template"); dl_b.clicked.connect(self.dl_template); layout.addWidget(dl_b); layout.addStretch()

    def browse(self):
        path, _ = QFileDialog.getOpenFileName(self, "Select Excel", "", "Excel Files (*.xlsx)")
        if path: self.path_edit.setText(path)

    def dl_template(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Template", "Consumer_Template.xlsx", "Excel Files (*.xlsx)")
        if path:
            with open(path, "wb") as f: f.write(self.utils.get_excel_template_consumers())
            QMessageBox.information(self, "Saved", f"Template saved to {path}")

    def run_import(self):
        path = self.path_edit.text()
        if not path: return
        self.progress.setVisible(True); self.progress.setRange(0, 0)
        def run():
            ws = openpyxl.load_workbook(path, data_only=True).active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h).strip() for h in rows[0]]
            data = []
            for r in rows[1:]:
                if not any(r): continue
                payload = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
                data.append(payload)
            return self.fc.bulk_create_consumers(data, self.admin_ctx["name"])
        def done(res):
            self.progress.setVisible(False)
            QMessageBox.information(self, "Import Complete", f"Imported {res['success']} consumers. Errors: {len(res['errors'])}")
        self.utils.run_in_thread(run, callback=done, error_callback=lambda e: QMessageBox.critical(self, "Error", str(e)))

class ExportTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self); layout.addWidget(QLabel("<b>Export CSD to Excel</b>"))
        self.zone_cb = QComboBox(); self.zone_cb.addItems(["All"] + [str(z) for z in self.utils.ZONE_RANGE]); layout.addWidget(QLabel("Zone:")); layout.addWidget(self.zone_cb)
        export_b = QPushButton("📤 Export Now"); export_b.clicked.connect(self.run_export); layout.addWidget(export_b); layout.addStretch()

    def run_export(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Export", "Consumers_Export.xlsx", "Excel Files (*.xlsx)")
        if not path: return
        z = self.zone_cb.currentText()
        def run():
            filters = {}
            if z != "All": filters["zone"] = int(z)
            consumers = self.fc.list_consumers(filters)
            wb = openpyxl.Workbook(); ws = wb.active
            headers = ["cin_no", "name", "zone", "category", "water_supply_type", "has_sewer_connection", "outstanding_balance"]
            ws.append(headers)
            for c in consumers: ws.append([c.get(h) for h in headers])
            wb.save(path)
        self.utils.run_in_thread(run, callback=lambda _: QMessageBox.information(self, "Done", "Export finished."))

class MeterReplacementTab(QWidget):
    def __init__(self, parent, fc, utils, be, admin_ctx):
        super().__init__(parent)
        self.fc = fc; self.utils = utils; self.admin_ctx = admin_ctx; self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        self.cin_input = QLineEdit(); self.cin_input.setPlaceholderText("Enter CIN..."); layout.addWidget(self.cin_input)
        search_b = QPushButton("🔍 Find Consumer"); search_b.clicked.connect(self.search); layout.addWidget(search_b)
        self.info = QLabel("No consumer selected."); layout.addWidget(self.info)
        self.form = QGroupBox("Replacement Form"); self.form.setEnabled(False); lay = QFormLayout(self.form)
        self.f_serial = QLineEdit(); self.f_kl = QDoubleSpinBox(); self.f_kl.setRange(0, 999999)
        lay.addRow("New Serial:", self.f_serial); lay.addRow("New KL:", self.f_kl); layout.addWidget(self.form)
        save_b = QPushButton("⚙️ Record Replacement"); save_b.clicked.connect(self.save); layout.addWidget(save_b); layout.addStretch()

    def search(self):
        def done(c):
            if not c: return QMessageBox.warning(self, "Error", "Not found.")
            self.c = c; self.info.setText(f"Found: {c['name']} (Current: {c.get('meter_serial_no')})"); self.form.setEnabled(True)
        self.utils.run_in_thread(lambda: self.fc.get_consumer(self.cin_input.text().strip()), callback=done)

    def save(self):
        d = {"cin_no": self.c["cin_no"], "old_serial": self.c.get("meter_serial_no"), "new_serial": self.f_serial.text().strip(),
             "replacement_date": datetime.now().strftime("%d-%m-%Y"), "new_initial_reading_kl": self.f_kl.value(), "admin_name": self.admin_ctx["name"]}
        self.utils.run_in_thread(lambda: self.fc.record_meter_replacement(**d), callback=lambda lid: QMessageBox.information(self, "Success", f"Log: {lid}"))
